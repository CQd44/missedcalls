"""
Missed Call Uploader and Verifier
Agents use this to verify calls they've returned.
Admins use it to upload files to populate those lists.
Port 13798
"""

import csv
import json
import math
import os
import socket
import toml
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Generator, List, Optional, Tuple

import aiofiles
import openpyxl
import psycopg2
import psycopg2.extras
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
CONFIG = toml.load("./config.toml")
CONNECT_STR = (
    f"dbname={CONFIG['credentials']['dbname']} "
    f"user={CONFIG['credentials']['username']} "
    f"password={CONFIG['credentials']['password']} "
    f"host={CONFIG['credentials']['host']}"
)

HEADERS = ["Queue Name", "Call Time", "Contact Disposition", "Phone Number"]
TEMP_DIR = "temp_files"
os.makedirs(TEMP_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------
def get_connection():
    """Return a new psycopg2 connection."""
    return psycopg2.connect(CONNECT_STR)

def get_hostname(ip_address: str) -> str:
    """Resolve an IP address to a hostname via reverse DNS lookup."""
    try:
        return socket.gethostbyaddr(ip_address)[0]
    except socket.herror:
        return "No hostname found"


# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------
@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        print("Started")
    except Exception as e:
        print(f"[startup] DB init error: {e}")
    yield


app = FastAPI(lifespan=lifespan)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.get("/")
async def home(request: Request):
    """Queue selection page – dropdown of queues with unreturned calls today."""
    # Reset the easter-egg click counter when visiting the homepage
    ip = request.client.host if request.client else "unknown"
    _logo_clicks[ip] = 0
    con = get_connection()
    cur = con.cursor()
    cur.execute(
        "SELECT DISTINCT(queue) FROM missedcalls WHERE (returned = FALSE AND date(time) = CURRENT_DATE);"
    )
    results = cur.fetchall()
    cur.close()
    con.close()
    queues = sorted([item[0] for item in results])
    return templates.TemplateResponse(
        request,
        "home.html",
        {
            "queues": queues,
        },
    )


@app.get("/calls")
async def calls(request: Request, queue: str = ""):
    """Show unreturned calls for a specific queue (today only)."""
    if not queue:
        return RedirectResponse(url="/", status_code=302)

    now = datetime.now()
    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute(
                """
                SELECT queue, time, phone, dialed
                FROM missedcalls
                WHERE queue = %s
                AND returned = FALSE
                AND date(time) = CURRENT_DATE
                ORDER BY phone;
                """,
                (queue,),
            )
            rows = cur.fetchall()

    call_list = []
    for row in rows:
        call_list.append({
            "queue": row[0],
            "time": row[1],
            "phone": row[2],
            "dialed": row[3] or "",
        })

    total = len(call_list)

    # Weekly performance for this queue (same week window as the original tool)
    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute(
                """
                SELECT
                    COUNT(*) AS missed_calls,
                    COUNT(*) FILTER (WHERE returned = TRUE) AS returned_calls
                FROM missedcalls
                WHERE queue = %s
                AND date(time) >= date_trunc('week', CURRENT_DATE)
                AND date(time) <  date_trunc('week', CURRENT_DATE) + INTERVAL '7 days';
                """,
                (queue,),
            )
            weekly = cur.fetchone()

    weekly_missed = weekly[0] if weekly else 0
    weekly_returned = weekly[1] if weekly else 0
    weekly_rate = math.ceil(100 * (weekly_returned / weekly_missed)) if weekly_missed else 0

    # Gauge needle: 0% => -90deg, 100% => +90deg
    needle_angle = -90 + (weekly_rate * 1.8)

    return templates.TemplateResponse(
        request,
        "calls.html",
        {
            "queue_name": queue,
            "calls": call_list,
            "total_calls": total,
            "weekly_missed": weekly_missed,
            "weekly_returned": weekly_returned,
            "weekly_rate": weekly_rate,
            "current_time": now.strftime("%I:%M %p"),
            "current_date": now.strftime("%m/%d/%Y"),
            "needle_angle": needle_angle,
        },
    )


@app.get("/calls/stats")
async def calls_stats(request: Request, queue: str = "", period: str = "week"):
    """Return JSON stats (abandoned, returned, rate) for a queue over a date period."""
    if not queue:
        raise HTTPException(status_code=400, detail="Queue is required.")

    period_clauses = {
        "day": "date(time) = CURRENT_DATE",
        "week": "date(time) >= date_trunc('week', CURRENT_DATE) AND date(time) < date_trunc('week', CURRENT_DATE) + INTERVAL '7 days'",
        "month": "date(time) >= date_trunc('month', CURRENT_DATE) AND date(time) < date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'",
    }
    clause = period_clauses.get(period)
    if not clause:
        raise HTTPException(status_code=400, detail="Invalid period. Use day, week, or month.")

    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                    COUNT(*) AS missed_calls,
                    COUNT(*) FILTER (WHERE returned = TRUE) AS returned_calls
                FROM missedcalls
                WHERE queue = %s
                AND {clause};
                """,
                (queue,),
            )
            row = cur.fetchone()

    abandoned = row[0] if row else 0
    returned = row[1] if row else 0
    rate = math.ceil(100 * (returned / abandoned)) if abandoned else 0

    return JSONResponse({
        "abandoned": abandoned,
        "returned": returned,
        "rate": rate,
    })


@app.post("/clearcalls")
async def clear_calls(request: Request, calls_json: str = Form(""), queue: str = Form("")):
    """Mark selected calls as returned by matching (queue, time, phone, dialed)."""
    if not calls_json:
        raise HTTPException(status_code=400, detail="No calls provided.")

    try:
        calls = json.loads(calls_json)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid calls data.")

    if not isinstance(calls, list) or len(calls) == 0:
        raise HTTPException(status_code=400, detail="No valid calls to mark.")

    now = datetime.now()
    client_ip = request.client.host # type: ignore
    hostname = get_hostname(client_ip)

    with get_connection() as con:
        with con.cursor() as cur:
            for call in calls:
                call_time = call.get("time", "")
                phone = call.get("phone", "")
                dialed = call.get("dialed", "")

                try:
                    time_val = datetime.strptime(call_time, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue

                cur.execute(
                    """
                    UPDATE missedcalls
                    SET returned = TRUE,
                        returned_on = %s,
                        ip_address = %s,
                        hostname = %s
                    WHERE queue = %s
                      AND time = %s
                      AND phone = %s
                      AND COALESCE(dialed::text, '') = %s
                      AND returned = FALSE;
                    """,
                    (now, client_ip, hostname, queue, time_val, phone, dialed),
                )
        con.commit()

    # Check if any unreturned calls remain for THIS queue today
    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute(
                """
                SELECT COUNT(*) FROM missedcalls
                WHERE queue = %s AND returned = FALSE AND date(time) = CURRENT_DATE;
                """,
                (queue,),
            )
            remaining = cur.fetchone()[0] # type: ignore

    if remaining == 0:
        return RedirectResponse(url="/", status_code=303)
    return RedirectResponse(url=f"/calls?queue={queue}", status_code=303)


@app.get("/dashboard")
async def dashboard(request: Request):
    """Dashboard with all-queue stats and a ring chart."""
    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute(
                """
                SELECT queue,
                       COUNT(*) AS total,
                       COUNT(*) FILTER (WHERE time >= NOW() - INTERVAL '7 days') AS this_week,
                       COUNT(*) FILTER (WHERE time >= NOW() - INTERVAL '30 days') AS last_30,
                       COUNT(*) FILTER (WHERE returned = TRUE) AS returned
                FROM missedcalls
                GROUP BY queue
                ORDER BY queue;
                """
            )
            rows = cur.fetchall()

            cur.execute("SELECT COUNT(*), COUNT(*) FILTER (WHERE returned = TRUE) FROM missedcalls;")
            overall = cur.fetchone()

    overall_missed = overall[0] if overall else 0
    overall_returned = overall[1] if overall else 0
    overall_rate = round((overall_returned / overall_missed) * 100) if overall_missed else 0

    dashboard_data = []
    for r in rows:
        rate = round((r[4] / r[1]) * 100) if r[1] else 0
        dashboard_data.append([r[0], r[1], r[2], r[3], r[4], rate])

    return templates.TemplateResponse(
        request,
        "dashboard.html",
        {
            "dashboard_data": dashboard_data,
            "overall_missed": overall_missed,
            "overall_returned": overall_returned,
            "overall_rate": overall_rate,
        },
    )

@app.get("/dashboard/stats")
async def dashboard_stats(request: Request, period: str = "week"):
    """Return JSON stats for all queues over a given period (day/week/month)."""
    period_clauses = {
        "day": "date(time) = CURRENT_DATE",
        "week": "date(time) >= date_trunc('week', CURRENT_DATE) AND date(time) < date_trunc('week', CURRENT_DATE) + INTERVAL '7 days'",
        "month": "date(time) >= date_trunc('month', CURRENT_DATE) AND date(time) < date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'",
    }
    clause = period_clauses.get(period)
    if not clause:
        raise HTTPException(status_code=400, detail="Invalid period. Use day, week, or month.")

    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute(
                f"""
                SELECT queue,
                       COUNT(*) AS total,
                       COUNT(*) FILTER (WHERE returned = TRUE) AS returned
                FROM missedcalls
                WHERE {clause}
                GROUP BY queue
                ORDER BY queue;
                """,
            )
            rows = cur.fetchall()

            cur.execute(
                f"""
                SELECT COUNT(*), COUNT(*) FILTER (WHERE returned = TRUE)
                FROM missedcalls
                WHERE {clause};
                """
            )
            overall = cur.fetchone()

    queues = []
    for r in rows:
        rate = round((r[2] / r[1]) * 100) if r[1] else 0
        queues.append({"name": r[0], "total": r[1], "returned": r[2], "rate": rate})

    overall_total = overall[0] if overall else 0
    overall_returned = overall[1] if overall else 0
    overall_rate = round((overall_returned / overall_total) * 100) if overall_total else 0

    return JSONResponse({
        "period": period,
        "overall_total": overall_total,
        "overall_returned": overall_returned,
        "overall_rate": overall_rate,
        "queues": queues,
    })

@app.post("/upload")
async def process_upload(request: Request, csv_file: UploadFile = File(...)):
    """Process an uploaded CSV or XLSX file and insert rows into the DB."""
    if not csv_file.filename:
        raise HTTPException(status_code=400, detail="No file provided.")

    file = csv_file
    input_file = ''

    if file.filename[-1] == 'x' and file.filename[:5] != 'Agent':
        input_file = 'temp_files\\temp_file.csv'
        if not os.path.exists(f'temp_files\\{file.filename}'):
            try:
                contents = await file.read()
                async with aiofiles.open(f"temp_files\\{file.filename}", 'wb') as f:
                    await f.write(contents)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f'Something went wrong. Tell Clay! {e}')
            finally:
                await file.close()
            wb = openpyxl.load_workbook(filename=f'temp_files\\{file.filename}', data_only=True)
            sheet = wb.worksheets[0]
            reader = sheet.iter_rows(values_only=True)
            first_row_skipped = False
            input_rows: list = []
            for row in reader:
                if row[0] == None:
                    if not first_row_skipped:
                        first_row_skipped = True
                    else:
                        break
                    continue
                elif row not in input_rows:
                    input_rows.append(row)

            with open(input_file, 'w', newline='') as temp:
                writer = csv.writer(temp)
                for row in input_rows:
                    writer.writerow(row)
    elif file.filename[-1] == 'x' and file.filename[:5] == 'Agent':
        print("Urology file detected")
        if not os.path.exists(f'temp_files\\{file.filename}'):
            try:
                contents = await file.read()
                async with aiofiles.open(f"temp_files\\{file.filename}", 'wb') as f:
                    await f.write(contents)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f'Something went wrong. Tell Clay! {e}')
            finally:
                await file.close()
        handle_xlsx(f"temp_files\\{file.filename}")
        input_file = "temp_files\\urology_output.csv"

    if file.filename[-1] == 'v':
        if not os.path.exists(f'temp_files\\{file.filename}'):
            try:
                contents = await file.read()
                async with aiofiles.open(f"temp_files\\{file.filename}", 'wb') as f:
                    await f.write(contents)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f'Something went wrong. Tell Clay! {e}')
            finally:
                await file.close()
        input_file = f"temp_files\\{file.filename}"

    def row_generator() -> Generator:
        with open(input_file, 'r', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                try:
                    yield row
                except Exception as e:
                    print("Invalid row.\n", e)

    row_gen = row_generator()
    con = psycopg2.connect(CONNECT_STR)
    cur = con.cursor()
    QUERY = "INSERT into missedcalls (queue, time, phone, dialed) VALUES (%s, %s, %s, %s) ON CONFLICT (queue, time, phone) DO NOTHING;"
    to_insert = []
    for row in row_gen:
        if row['Contact Disposition'] in {'1', '1.0'}:
            try:
                to_insert.append((row['Queue Name'], row['Call Time'], int(row['Phone Number']), int(row['Number Dialed'])))
            except (ValueError, TypeError):
                print("Skipped row (invalid phone/dialed):", row)
    if to_insert:
        cur.executemany(QUERY, to_insert)
        rows_added = cur.rowcount
    else:
        rows_added = 0

    cur.close()
    con.commit()
    con.close()

    try:
        files = os.listdir("temp_files")
        for x in files:
            os.remove(f'temp_files\\{x}')
    except:
        pass

    if rows_added == 0:
        return HTMLResponse(content="File uploaded and processed successfully. No new calls were added to the database.")
    else:
        return HTMLResponse(content=f"File uploaded and processed successfully. {rows_added} new calls were added to the database.")

def handle_xlsx(input):
    calls_presented: int = 0
    calls_handled: int = 0
    presented_dict: dict = {}
    handled_dict: dict = {}
    wb = openpyxl.load_workbook(filename=f'{input}', data_only=True)
    sheet = wb.worksheets[0]
    reader = sheet.iter_rows(values_only=True)
    first_row_skipped = False
    input_rows: list = []
    for row in reader:
        if row[0] == None:
            if not first_row_skipped:
                first_row_skipped = True
            else:
                break
            continue
        elif row not in input_rows:
            input_rows.append(row)

    with open('temp_files\\temp_file.csv', 'w', newline='') as temp:
        writer = csv.writer(temp)
        for row in input_rows:
            writer.writerow(row)
    previous_row = {}
    with open('temp_files\\temp_file.csv', "r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            if row['Extension'] == row['Call ANI'] and row['Called Number'] == '21898':
                calls_presented += 1
                presented_dict[row['Call Start Time']] = previous_row['Call ANI']
            if row['Extension'] != row['Call ANI'] and row['Called Number'] == '21898':
                calls_handled += 1
                handled_dict[row['Call Start Time']] = row['Call ANI']
            previous_row = row

    presented_numbers: list[str] = list(presented_dict.values())
    handled_numbers: list[str] = list(handled_dict.values())
    abandoned_numbers: dict[str, str] = {}

    for number in presented_numbers:
        call_time = ''
        if number not in handled_numbers and number not in abandoned_numbers:
            for time in presented_dict.keys():
                if number == presented_dict[time]:
                    call_time = time
            abandoned_numbers[call_time] = number

    with open("temp_files\\urology_output.csv", 'w', newline='') as output:
        writer = csv.writer(output)
        header = ['Queue Name', 'Call Time', 'Phone Number', 'Contact Disposition']
        writer.writerow(header)
        for time in abandoned_numbers.keys():
            writer.writerow(['URO_SCHL_CSQ', time, abandoned_numbers[time], '1'])


def _fetch_report_data(from_date: datetime, to_date: datetime):
    """Query the DB and return structured report data grouped by queue."""
    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute(
                """
                SELECT queue, time, phone, returned, returned_on, ip_address, hostname
                FROM missedcalls
                WHERE DATE(time) >= %s AND DATE(time) <= %s
                ORDER BY queue, time;
                """,
                (from_date, to_date),
            )
            results = cur.fetchall()

    # Group by queue
    from collections import defaultdict
    queues = defaultdict(list)
    for row in results:
        queues[row[0]].append({
            "time": row[1],
            "phone": row[2],
            "returned": row[3],
            "returned_on": row[4],
            "ip_address": str(row[5]) if row[5] else "",
            "hostname": row[6] if row[6] else "",
        })

    # Build per-queue summaries
    queue_data = []
    overall_total = 0
    overall_returned = 0

    for qname in sorted(queues.keys()):
        calls = queues[qname]
        total = len(calls)
        returned = sum(1 for c in calls if c["returned"])
        not_returned = total - returned
        rate = round((returned / total) * 100) if total else 0

        # Daily breakdown
        daily = defaultdict(int)
        for c in calls:
            daily[c["time"].strftime("%m/%d")] += 1
        daily_sorted = sorted(daily.items())

        queue_data.append({
            "name": qname,
            "total": total,
            "returned": returned,
            "not_returned": not_returned,
            "rate": rate,
            "daily": daily_sorted,
            "calls": calls,
        })

        overall_total += total
        overall_returned += returned

    overall_not_returned = overall_total - overall_returned
    overall_rate = round((overall_returned / overall_total) * 100) if overall_total else 0

    return {
        "overall_total": overall_total,
        "overall_returned": overall_returned,
        "overall_not_returned": overall_not_returned,
        "overall_rate": overall_rate,
        "queues": queue_data,
    }


@app.post("/report")
async def report_form(request: Request, month: int = Form(...), day: int = Form(...), year: int = Form(...)):
    """Handle the date form submission from the home page."""
    try:
        from_date = datetime(year, month, day)
    except ValueError:
        return templates.TemplateResponse(
            request,
            "report_error.html",
            {"error_message": "Invalid date selected."},
        )

    to_date = datetime.now()
    data = _fetch_report_data(from_date, to_date)

    return templates.TemplateResponse(
        request,
        "report.html",
        {
            "date_from": from_date.strftime("%Y-%m-%d"),
            "date_to": to_date.strftime("%Y-%m-%d"),
            "data": data,
        },
    )

@app.get("/report")
async def report_page(request: Request, date_from: str = "", date_to: str = ""):
    """Render the interactive HTML report page with per-queue accordions."""
    try:
        from_date = datetime.strptime(date_from, "%Y-%m-%d")
        to_date = datetime.strptime(date_to, "%Y-%m-%d")
    except (ValueError, TypeError):
        return templates.TemplateResponse(request, "report_error.html")

    if to_date < from_date:
        return templates.TemplateResponse(request, "report_error.html")

    data = _fetch_report_data(from_date, to_date)

    return templates.TemplateResponse(
        request,
        "report.html",
        {
            "date_from": date_from,
            "date_to": date_to,
            "data": data,
        },
    )

@app.get("/report/download_csv")
async def download_csv(date_from: str = "", date_to: str = "", queues: str = ""):
    """Download the report as a flat CSV. Optionally filter by comma-separated queue names."""
    try:
        from_date = datetime.strptime(date_from, "%Y-%m-%d")
        to_date = datetime.strptime(date_to, "%Y-%m-%d")
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid dates.")

    queue_list = [q.strip() for q in queues.split(",") if q.strip()] if queues else None

    with get_connection() as con:
        with con.cursor() as cur:
            if queue_list:
                placeholders = ", ".join(["%s"] * len(queue_list))
                cur.execute(
                    f"""
                    SELECT queue, time, phone, returned, returned_on, ip_address, hostname
                    FROM missedcalls
                    WHERE DATE(time) >= %s AND DATE(time) <= %s AND queue IN ({placeholders})
                    ORDER BY time;
                    """,
                    [from_date, to_date] + queue_list,
                )
            else:
                cur.execute(
                    """
                    SELECT queue, time, phone, returned, returned_on, ip_address, hostname
                    FROM missedcalls
                    WHERE DATE(time) >= %s AND DATE(time) <= %s
                    ORDER BY time;
                    """,
                    (from_date, to_date),
                )
            results = cur.fetchall()

    with open("callbacks_report.csv", "w", newline="") as output:
        writer = csv.writer(output)
        writer.writerow(["Queue", "Date and Time of Call", "Phone Number", "Returned", "Returned On", "IP Address", "PC Name"])
        for row in results:
            writer.writerow([
                row[0],
                row[1].strftime("%m/%d/%y %I:%M:%S %p") if row[1] else "",
                row[2],
                "Yes" if row[3] else "No",
                row[4].strftime("%m/%d/%y %I:%M:%S %p") if row[4] else "",
                str(row[5]) if row[5] else "",
                row[6] if row[6] else "",
            ])

    return FileResponse(
        path="callbacks_report.csv",
        media_type="text/csv",
        filename="callbacks_report.csv",
    )

@app.get("/report/download_excel")
async def download_excel(date_from: str = "", date_to: str = "", queues: str = ""):
    """Download the report as a multi-sheet Excel file. Optionally filter by comma-separated queue names."""
    try:
        from_date = datetime.strptime(date_from, "%Y-%m-%d")
        to_date = datetime.strptime(date_to, "%Y-%m-%d")
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid dates.")

    queue_list = [q.strip() for q in queues.split(",") if q.strip()] if queues else None
    data = _fetch_report_data(from_date, to_date)

    # Filter queues if specified
    if queue_list:
        data["queues"] = [q for q in data["queues"] if q["name"] in queue_list]
        data["overall_total"] = sum(q["total"] for q in data["queues"])
        data["overall_returned"] = sum(q["returned"] for q in data["queues"])
        data["overall_not_returned"] = data["overall_total"] - data["overall_returned"]
        data["overall_rate"] = round((data["overall_returned"] / data["overall_total"]) * 100) if data["overall_total"] else 0

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary" # type: ignore
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="00A9A7", end_color="00A9A7", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws["A1"] = "Abandoned Call Report" # type: ignore
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Date Range: {date_from} to {date_to}"
    ws["A3"] = f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"

    row = 5
    ws.cell(row=row, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total Calls")
    ws.cell(row=row, column=2, value=data["overall_total"])
    row += 1
    ws.cell(row=row, column=1, value="Returned")
    ws.cell(row=row, column=2, value=data["overall_returned"])
    row += 1
    ws.cell(row=row, column=1, value="Not Returned")
    ws.cell(row=row, column=2, value=data["overall_not_returned"])
    row += 1
    ws.cell(row=row, column=1, value="Return Rate")
    ws.cell(row=row, column=2, value=f"{data['overall_rate']}%")

    row += 2
    ws.cell(row=row, column=1, value="Queue").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Returned").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Not Returned").font = Font(bold=True)
    ws.cell(row=row, column=5, value="Rate %").font = Font(bold=True)
    row += 1
    for q in data["queues"]:
        ws.cell(row=row, column=1, value=q["name"])
        ws.cell(row=row, column=2, value=q["total"])
        ws.cell(row=row, column=3, value=q["returned"])
        ws.cell(row=row, column=4, value=q["not_returned"])
        ws.cell(row=row, column=5, value=f"{q['rate']}%")
        row += 1

    # --- Per-queue sheets ---
    for q in data["queues"]:
        sheet_name = q["name"][:31].replace("/", "-").replace("\\", "-").replace("[", "(").replace("]", ")").replace("*", "").replace("?", "").replace(":", "-")
        ws_q = wb.create_sheet(title=sheet_name)

        ws_q["A1"] = f"Queue: {q['name']}"
        ws_q["A1"].font = Font(bold=True, size=13)
        ws_q["A2"] = f"Total: {q['total']}  |  Returned: {q['returned']}  |  Not Returned: {q['not_returned']}  |  Rate: {q['rate']}%"

        headers = ["Call Time", "Phone Number", "Returned", "Returned On", "IP Address", "PC Name"]
        for col, h in enumerate(headers, 1):
            cell = ws_q.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, call in enumerate(q["calls"], start=5):
            ws_q.cell(row=i, column=1, value=call["time"].strftime("%m/%d/%Y %I:%M:%S %p"))
            ws_q.cell(row=i, column=2, value=call["phone"])
            ret_cell = ws_q.cell(row=i, column=3, value="Yes" if call["returned"] else "No")
            ret_cell.fill = green_fill if call["returned"] else red_fill
            ws_q.cell(row=i, column=4, value=call["returned_on"].strftime("%m/%d/%Y %I:%M:%S %p") if call["returned_on"] else "")
            ws_q.cell(row=i, column=5, value=call["ip_address"])
            ws_q.cell(row=i, column=6, value=call["hostname"])

        for col in range(1, 7):
            ws_q.column_dimensions[get_column_letter(col)].width = 18

    excel_path = "callbacks_report.xlsx"
    wb.save(excel_path)

    return FileResponse(
        path=excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="callbacks_report.xlsx",
    )

@app.get("/report_error")
async def report_error(request: Request):
    return templates.TemplateResponse(request, "report_error.html")

# ---------------------------------------------------------------------------
# Easter Egg — Credits
# ---------------------------------------------------------------------------
_logo_clicks: dict[str, int] = {}

@app.post("/logo_click")
async def logo_click(request: Request):
    """Increment the secret click counter for this client."""
    ip = request.client.host if request.client else "unknown"
    _logo_clicks[ip] = _logo_clicks.get(ip, 0) + 1
    return JSONResponse({"ok": True})

@app.get("/credits")
async def credits(request: Request):
    """Secret credits page — only accessible after 5 head clicks."""
    ip = request.client.host if request.client else "unknown"
    count = _logo_clicks.get(ip, 0)
    if count < 5:
        raise HTTPException(status_code=404, detail="Not found.")
    return templates.TemplateResponse(
        request,
        "credits.html",
    )

@app.get("/leaderboard")
async def get_leaderboard(request: Request):
    """Return the top 10 leaderboard entries."""
    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute("SELECT name, score FROM leaderboard ORDER BY score DESC LIMIT 10;")
            rows = cur.fetchall()
    return JSONResponse([{"name": r[0], "score": r[1]} for r in rows])

@app.post("/leaderboard")
async def add_leaderboard(request: Request, name: str = Form(...), score: int = Form(...)):
    """Insert a new leaderboard entry."""
    if not name.strip():
        raise HTTPException(status_code=400, detail="Name is required.")
    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute("INSERT INTO leaderboard (name, score) VALUES (%s, %s);", (name.strip(), score))
        con.commit()
    return JSONResponse({"ok": True})

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=13798, reload=False)
